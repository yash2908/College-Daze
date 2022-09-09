from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from .models import *
import xlrd
import os
from django.conf import settings
from openpyxl import load_workbook
from io import BytesIO
from payback.forms import *
# from django.conf.settings import PROJECT_ROOT
from payback.models import *

# Create your views here.
def opening(request):
    return render(request, 'opening.html')

@login_required(login_url='/login')
def firstyear(request):
    return render(request, 'firstyear.html')

@login_required(login_url='/login')
def kenken(request):
    return render(request, 'kenken.html')

@login_required(login_url='/login')
def slidepuzzle(request):
    return render(request, 'slidingpuzzle.html')

@login_required(login_url='/login')
def slider(request):
    return render(request, 'slider.html')

@login_required(login_url='/login')
def tangram(request):
    return render(request, 'tangram.html')

@login_required(login_url='/login')
def secondyear(request):
    technoplayer1 = Technoplayer1.objects.filter(user=request.user).first()
    loan1 = technoplayer1.loan1
    connection1 = technoplayer1.connection1
    happiness1 = technoplayer1.happiness1
    focus1 = technoplayer1.focus1
    return render(request, 'secondyear.html',
                  {'loan1': loan1, 'connection1': connection1, 'happiness1': happiness1, 'focus1': focus1})

@login_required(login_url='/login')
def thirdyear(request):
    technoplayer2 = Technoplayer2.objects.filter(user=request.user).first()
    loan2 = technoplayer2.loan2
    connection2 = technoplayer2.connection2
    happiness2 = technoplayer2.happiness2
    focus2 = technoplayer2.focus2
    return render(request, 'thirdyear.html',
                  {'loan2': loan2, 'connection2': connection2, 'happiness2': happiness2, 'focus2': focus2})

@login_required(login_url='/login')
def fourthyear(request):
    technoplayer3 = Technoplayer3.objects.filter(user=request.user).first()
    loan3 = technoplayer3.loan3
    connection3 = technoplayer3.connection3
    happiness3 = technoplayer3.happiness3
    focus3 = technoplayer3.focus3
    return render(request, 'fourthyear.html',
                  {'loan3': loan3, 'connection3': connection3, 'happiness3': happiness3, 'focus3': focus3})

@login_required(login_url='/login')
def graduation(request):
    return render(request, 'graduation.html')

@login_required(login_url='/login')
def game(request):
    return render(request, 'game.html')

@login_required(login_url='/login')
def mastermind(request):
    return render(request, 'mastermind.html')

@login_required(login_url='/login')
def crossword(request):
    return render(request, 'crossword_begining.html')

@login_required(login_url='/login')
def mysteryroom(request):
    return render(request, 'mystery_room2.html')


def firstyear_submission(request):
    technoplayer1 = Technoplayer1.objects.filter(user=request.user).first()
    if request.method == "POST":
        focus = request.POST.get('focus')
        happiness = request.POST.get('happiness')
        connection = request.POST.get('connection')
        loan = request.POST.get('loan')
        print(focus)
        print(loan)
        if technoplayer1 is None:
            print(1)
            technoplayer1 = Technoplayer1()
            technoplayer1.user = request.user
            technoplayer1.happiness1 = happiness
            technoplayer1.connection1 = connection
            technoplayer1.focus1 = focus
            technoplayer1.loan1 = loan
            technoplayer1.save()

        data = "Save Successfully"
        return JsonResponse(data, safe=False)
    
    return HttpResponse("get method")


def techo_login(request):
    if request.user.is_authenticated:
        return redirect(request.GET.get('next', '/firstyear'))

    if request.method == "POST":
        technouser = User.objects.filter(roll_no=request.POST['roll_no']).first()
        if technouser is None:
            return render(request, 'login.html', {"messages": [["text-danger", "Roll Number Not Found."]]})
        username = technouser.username
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            print("Used logged in!")
            return redirect(request.GET.get('next', '/firstyear'))
        else:
            return render(request, 'login.html', {"messages": [["text-danger", "Invalid Credentials."]]})
    return render(request, 'login.html', )


def logout_view(request):
    logout(request)
    return redirect('/')


def secondyear_submission(request):
    technoplayer2 = Technoplayer2.objects.filter(user=request.user).first()
    if request.method == "POST":
        focus = request.POST.get('focus2')
        print(focus)
        happiness = request.POST.get('happiness2')
        connection = request.POST.get('connection2')
        loan = request.POST.get('loan2')

        if technoplayer2 is  None:
            technoplayer2 = Technoplayer2()
            technoplayer2.user = request.user
            technoplayer2.happiness2 = happiness
            technoplayer2.connection2 = connection
            technoplayer2.focus2 = focus
            technoplayer2.loan2 = loan
            technoplayer2.save()

        data = "Save Successfully"
        return JsonResponse(data, safe=False)

    return HttpResponse("get method")


def thirdyear_submission(request):
    technoplayer3 = Technoplayer3.objects.filter(user=request.user).first()
    if request.method == "POST":
        focus = request.POST.get('focus3')
        print(focus)
        happiness = request.POST.get('happiness3')
        connection = request.POST.get('connection3')
        loan = request.POST.get('loan3')

        if technoplayer3 is None:
            technoplayer3 = Technoplayer3()
            technoplayer3.user = request.user
            technoplayer3.happiness3 = happiness
            technoplayer3.connection3 = connection
            technoplayer3.focus3 = focus
            technoplayer3.loan3 = loan
            technoplayer3.save()

        data = "Save Successfully"
        return JsonResponse(data, safe=False)

    return HttpResponse("get method")


def fourthyear_submission(request):
    technoplayer4 = Technoplayer4.objects.filter(user=request.user).first()
    if request.method == "POST":
        focus = request.POST.get('focus4')
        print(focus)
        happiness = request.POST.get('happiness4')
        connection = request.POST.get('connection4')
        loan = request.POST.get('loan4')
        if technoplayer4 is None:
            technoplayer4 = Technoplayer4()
            technoplayer4.user = request.user
            technoplayer4.happiness4 = happiness
            technoplayer4.connection4 = connection
            technoplayer4.focus4 = focus
            technoplayer4.loan4 = loan
            technoplayer4.save()

        data = "Save Successfully"
        return JsonResponse(data, safe=False)

    return HttpResponse("get method")


def kenken_submission(request):
    kenken_player = Kenkenplayer.objects.filter(user=request.user).first()
    if request.method == "POST":
        kenken_solver = request.POST.get('kenken_solver')
        print(kenken_solver)
        # print(1)
        if kenken_player is None:
            print(2)
            kenken_player = Kenkenplayer()
            kenken_player.user = request.user
            kenken_player.kenken_solver = kenken_solver
            kenken_player.save()

        data = "Save Successfully"
        return JsonResponse(data, safe=False)

    return HttpResponse("get method")


def mastermind_submission(request):
    mastermind_player = Mastermindplayer.objects.filter(user=request.user).first()
    if request.method == "POST":
        mastermind_solver = request.POST.get('mastermind_solver')
        print(mastermind_solver)
        # print(1)
        if mastermind_player is None:
            # print(2)
            mastermind_player = Mastermindplayer()
            mastermind_player.user = request.user
            mastermind_player.mastermind_solver = mastermind_solver
            mastermind_player.save()

        data = "Save Successfully"
        return JsonResponse(data, safe=False)

    return HttpResponse("get method")


def crossword_submission(request):
    crossword_player = Crosswordplayer.objects.filter(user=request.user).first()
    if request.method == "POST":
        crossword = request.POST.get('submittedCrossword')
        print(crossword)
        agesum = request.POST.get('is_agesum_solved')
        letter_sum = request.POST.get('is_lettersum_solved')
        puzzle_score = request.POST.get('puzzle_score')

        if crossword_player is  None:
            crossword_player = Crosswordplayer()
            crossword_player.user = request.user
            crossword_player.agesum = agesum
            crossword_player.letter_sum = letter_sum
            crossword_player.puzzle_score = puzzle_score
            crossword_player.submittedCrossword = crossword
            crossword_player.save()

        data = "Save Successfully"
        return JsonResponse(data, safe=False)

    return HttpResponse("get method")

def mysteryroom_submission(request):
    mysteryroom_player = Mysteryplayer.objects.filter(user=request.user).first()
    if request.method == "POST":
        jsonanswer = request.POST.get('JSONanswer')
        print(jsonanswer)
        if mysteryroom_player is None:
            mysteryroom_player = Mysteryplayer()
            mysteryroom_player.user = request.user
            mysteryroom_player.answers = jsonanswer
            mysteryroom_player.save()

        data = "Save Successfully"
        return JsonResponse(data, safe=False)

    return HttpResponse("get method")

def append_user(excel_data):
    wb = load_workbook(filename=BytesIO(excel_data))
    sheet = wb.active
    r = 114
    for i in  range(r):
        name_val = sheet.cell(row=i+2, column=1).value
        print(name_val)
        #name = name_val.split()
        contact = sheet.cell(row=i+2, column=2).value
        email = sheet.cell(row=i+2, column=3).value
        roll = str(sheet.cell(row=i+2, column=4).value)
        username = str(roll)
        password = str(sheet.cell(row=i+2, column=5).value)
        user = User.objects.create_user(username=username, email=email, password=password)
        user.first_name = name_val
        user.roll_no = roll
        user.contact = contact
        user.save()

def get_sheet(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file_in_memory = request.FILES['file'].read()
            # wb = load_workbook(filename=BytesIO(file_in_memory))
            append_user(file_in_memory)

            return HttpResponse("Read Successfully")


    else:
        form = UploadFileForm()
        return render(request, 'upload.html', {'form': form})
